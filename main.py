"""
DART(전자공시시스템) Open API를 이용한 재무제표 비교 분석기.

사용법:
    python main.py "삼성전자" "SK하이닉스" "LG전자" --years 2023 2022 2021 --api-key YOUR_KEY

API 키:
    https://opendart.fss.or.kr 에서 무료로 발급받을 수 있습니다.
    --api-key 옵션 대신 환경변수 DART_API_KEY 로 지정해도 됩니다.
"""

import argparse
import io
import os
import sys
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta

import requests
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_URL = "https://opendart.fss.or.kr/api"
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CORP_CODE_CACHE = os.path.join(SCRIPT_DIR, "corpCode.xml")
CORP_CODE_MAX_AGE_DAYS = 7

# 보고서 코드: 사업보고서(연간)
REPRT_CODE_ANNUAL = "11011"

# 비교할 핵심 계정과목: 표시명 -> 후보 계정명 목록 (보고 관행상 명칭이 다를 수 있어 여러 후보를 둠)
ACCOUNT_MAP = {
    "매출액": ["매출액", "수익(매출액)", "영업수익", "매출"],
    "영업이익": ["영업이익", "영업이익(손실)"],
    "당기순이익": ["당기순이익", "당기순이익(손실)", "분기순이익", "반기순이익"],
    "자산총계": ["자산총계"],
    "부채총계": ["부채총계"],
    "자본총계": ["자본총계"],
}


def get_api_key(cli_key):
    key = cli_key or os.environ.get("DART_API_KEY")
    if not key:
        sys.exit(
            "DART API 키가 필요합니다. --api-key 옵션을 사용하거나 "
            "환경변수 DART_API_KEY 를 설정하세요.\n"
            "키 발급: https://opendart.fss.or.kr"
        )
    return key


def download_corp_codes(api_key):
    """corpCode.xml을 다운로드해 로컬에 캐싱한다."""
    if os.path.exists(CORP_CODE_CACHE):
        age = datetime.now() - datetime.fromtimestamp(os.path.getmtime(CORP_CODE_CACHE))
        if age < timedelta(days=CORP_CODE_MAX_AGE_DAYS):
            return

    print("기업 코드 목록(corpCode) 다운로드 중...")
    resp = requests.get(f"{BASE_URL}/corpCode.xml", params={"crtfc_key": api_key}, timeout=30)
    resp.raise_for_status()

    try:
        with zipfile.ZipFile(io.BytesIO(resp.content)) as zf:
            name = zf.namelist()[0]
            data = zf.read(name)
    except zipfile.BadZipFile:
        # 인증 실패 등 오류 응답(XML/JSON 텍스트)이 온 경우
        sys.exit(f"corpCode 다운로드 실패. API 키를 확인하세요.\n응답: {resp.content[:300]}")

    with open(CORP_CODE_CACHE, "wb") as f:
        f.write(data)


def find_corp_code(company_name):
    """회사명으로 corp_code를 검색한다. 정확히 일치하는 상장사를 우선한다."""
    tree = ET.parse(CORP_CODE_CACHE)
    root = tree.getroot()

    exact_listed, exact_any, partial = [], [], []
    for corp in root.findall("list"):
        name = corp.findtext("corp_name", "").strip()
        code = corp.findtext("corp_code", "").strip()
        stock_code = corp.findtext("stock_code", "").strip()
        if name == company_name:
            (exact_listed if stock_code else exact_any).append((code, name, stock_code))
        elif company_name in name:
            partial.append((code, name, stock_code))

    candidates = exact_listed or exact_any or partial
    if not candidates:
        return None, None

    # 상장사(stock_code 존재)를 우선 선택
    candidates.sort(key=lambda x: x[2] == "", reverse=False)
    code, name, stock_code = candidates[0]
    return code, name


def fetch_financials(api_key, corp_code, year, fs_div="CFS"):
    """단일회사 전체 재무제표 조회. CFS(연결) 실패 시 OFS(별도)로 재시도."""
    for div in ([fs_div, "OFS"] if fs_div == "CFS" else [fs_div]):
        resp = requests.get(
            f"{BASE_URL}/fnlttSinglAcntAll.json",
            params={
                "crtfc_key": api_key,
                "corp_code": corp_code,
                "bsns_year": year,
                "reprt_code": REPRT_CODE_ANNUAL,
                "fs_div": div,
            },
            timeout=30,
        )
        resp.raise_for_status()
        data = resp.json()
        if data.get("status") == "000":
            return data["list"], div
    return [], fs_div


def extract_key_accounts(rows):
    """계정과목 리스트에서 핵심 지표만 추출."""
    result = {}
    for display_name, candidates in ACCOUNT_MAP.items():
        for row in rows:
            if row.get("account_nm", "").strip() in candidates:
                amount = row.get("thstrm_amount", "0").replace(",", "")
                try:
                    result[display_name] = int(amount)
                except ValueError:
                    result[display_name] = None
                break
        else:
            result[display_name] = None
    return result


def compute_ratios(metrics):
    revenue = metrics.get("매출액")
    op_income = metrics.get("영업이익")
    net_income = metrics.get("당기순이익")
    assets = metrics.get("자산총계")
    liabilities = metrics.get("부채총계")
    equity = metrics.get("자본총계")

    def pct(numerator, denominator):
        if numerator is None or not denominator:
            return None
        return round(numerator / denominator * 100, 2)

    return {
        "영업이익률(%)": pct(op_income, revenue),
        "순이익률(%)": pct(net_income, revenue),
        "부채비율(%)": pct(liabilities, equity),
        "ROE(%)": pct(net_income, equity),
        "ROA(%)": pct(net_income, assets),
    }


def build_workbook(all_data, output_path):
    """
    all_data: { company_name: { year: {metrics..., ratios...} } }
    """
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # 1) 요약 비교 시트: 행=지표, 열=회사-연도
        columns = []
        for company, by_year in all_data.items():
            for year in sorted(by_year.keys(), reverse=True):
                columns.append((company, str(year)))

        metric_rows = list(ACCOUNT_MAP.keys())
        ratio_rows = ["영업이익률(%)", "순이익률(%)", "부채비율(%)", "ROE(%)", "ROA(%)"]

        summary = pd.DataFrame(
            index=metric_rows + ratio_rows,
            columns=pd.MultiIndex.from_tuples(columns, names=["회사", "연도"]),
        )
        for company, year_str in columns:
            data = all_data[company][int(year_str)]
            for m in metric_rows:
                summary.loc[m, (company, year_str)] = data["metrics"].get(m)
            for r in ratio_rows:
                summary.loc[r, (company, year_str)] = data["ratios"].get(r)

        summary.to_excel(writer, sheet_name="요약비교")

        # 2) 회사별 상세 시트
        for company, by_year in all_data.items():
            rows = []
            for year in sorted(by_year.keys(), reverse=True):
                d = by_year[year]
                row = {"연도": year, "재무제표구분": d["fs_div"]}
                row.update(d["metrics"])
                row.update(d["ratios"])
                rows.append(row)
            df = pd.DataFrame(rows)
            sheet_name = company[:31]  # 엑셀 시트명 31자 제한
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    _format_workbook(output_path)


def _format_workbook(path):
    from openpyxl import load_workbook

    wb = load_workbook(path)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=2):
            for cell in row:
                if cell.value is not None:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
        for col_cells in ws.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max(length + 2, 12), 30)
        ws.freeze_panes = "B3" if ws.title == "요약비교" else "A2"

    wb.save(path)


def main():
    parser = argparse.ArgumentParser(description="DART 기반 3개년 재무제표 비교 분석기")
    parser.add_argument("companies", nargs=3, help="비교할 회사명 3개")
    parser.add_argument(
        "--years", nargs=3, type=int,
        default=[datetime.now().year - 1, datetime.now().year - 2, datetime.now().year - 3],
        help="조회할 연도 3개 (기본: 최근 3개년, 사업보고서 기준)",
    )
    parser.add_argument("--api-key", help="DART Open API 인증키")
    parser.add_argument("--fs-div", choices=["CFS", "OFS"], default="CFS",
                         help="CFS=연결재무제표(기본), OFS=별도재무제표")
    parser.add_argument("--output", default=None, help="출력 엑셀 파일 경로")
    args = parser.parse_args()

    api_key = get_api_key(args.api_key)
    download_corp_codes(api_key)

    all_data = {}
    for company_name in args.companies:
        corp_code, resolved_name = find_corp_code(company_name)
        if corp_code is None:
            print(f"'{company_name}' 에 해당하는 기업을 찾을 수 없습니다. 건너뜁니다.")
            continue
        if resolved_name != company_name:
            print(f"'{company_name}' -> '{resolved_name}' 로 매칭되었습니다.")

        by_year = {}
        for year in args.years:
            print(f"{resolved_name} {year}년 재무제표 조회 중...")
            rows, used_div = fetch_financials(api_key, corp_code, year, args.fs_div)
            if not rows:
                print(f"  경고: {resolved_name} {year}년 데이터를 찾을 수 없습니다.")
                continue
            metrics = extract_key_accounts(rows)
            ratios = compute_ratios(metrics)
            by_year[year] = {"metrics": metrics, "ratios": ratios, "fs_div": used_div}

        if by_year:
            all_data[resolved_name] = by_year

    if not all_data:
        sys.exit("조회된 데이터가 없어 엑셀 파일을 생성하지 않았습니다.")

    output_path = args.output or os.path.join(
        SCRIPT_DIR, f"재무분석_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    build_workbook(all_data, output_path)
    print(f"\n완료: {output_path}")


if __name__ == "__main__":
    main()
