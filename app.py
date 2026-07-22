"""재무분석기 - 로컬 웹앱

1차 필터: 산업군(네이버 증권)  ->  2차 필터: 회사명(네이버+DART)  ->  3차 필터: 계정과목(DART)
선택한 자료를 엑셀(.xlsx)로 다운로드.
"""

import io
import os
import re
import time
from collections import OrderedDict
from datetime import datetime

from flask import Flask, request, jsonify, send_file, render_template

import naver
import dart
import metrics
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


def _load_dotenv():
    """로컬 .env 파일이 있으면 환경변수로 읽어들인다(외부 의존성 없이)."""
    path = os.path.join(BASE_DIR, ".env")
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


_load_dotenv()

app = Flask(__name__, template_folder=os.path.join(BASE_DIR, "templates"))


def resolve_key(client_key):
    """클라이언트가 보낸 키를 우선하고, 없으면 서버 환경변수(DART_API_KEY)를 사용."""
    return (client_key or "").strip() or os.environ.get("DART_API_KEY", "").strip()

REPRT_NAMES = {
    "11011": "사업보고서(연간)",
    "11012": "반기보고서",
    "11013": "1분기보고서",
    "11014": "3분기보고서",
}
FS_NAMES = {"CFS": "연결재무제표", "OFS": "별도재무제표"}

# --- 플러스알파: 사업보고서 "내용" 카탈로그 -------------------------------

# 정기보고서 주요정보 (구조화 데이터): 그룹 > 세부항목(endpoint)
MAJOR_GROUPS = [
    {"group": "배당", "items": [
        {"id": "alotMatter", "name": "배당에 관한 사항"},
    ]},
    {"group": "주식·자본", "items": [
        {"id": "irdsSttus", "name": "증자(감자) 현황"},
        {"id": "tesstkAcqsDspsSttus", "name": "자기주식 취득·처분"},
    ]},
    {"group": "주주", "items": [
        {"id": "hyslrSttus", "name": "최대주주 현황"},
        {"id": "hyslrChgSttus", "name": "최대주주 변동현황"},
        {"id": "mrhlSttus", "name": "소액주주 현황"},
    ]},
    {"group": "임직원·보수", "items": [
        {"id": "exctvSttus", "name": "임원 현황"},
        {"id": "empSttus", "name": "직원 현황"},
        {"id": "hmvAuditAllSttus", "name": "이사·감사 전체 보수현황"},
        {"id": "hmvAuditIndvdlBySttus", "name": "개인별 보수(5억 이상)"},
    ]},
    {"group": "투자·채무", "items": [
        {"id": "otrCprInvstmntSttus", "name": "타법인 출자현황"},
        {"id": "cprndNrdmpBlce", "name": "회사채 미상환 잔액"},
        {"id": "srtpdPsndbtNrdmpBlce", "name": "단기사채 미상환 잔액"},
        {"id": "entrprsBilScritsNrdmpBlce", "name": "기업어음 미상환 잔액"},
        {"id": "detScritsIsuAcmslt", "name": "채무증권 발행실적"},
        {"id": "pssrpCptalUseDtls", "name": "공모자금 사용내역"},
        {"id": "prvsrpCptalUseDtls", "name": "사모자금 사용내역"},
    ]},
]
MAJOR_MAP = {it["id"]: it["name"] for g in MAJOR_GROUPS for it in g["items"]}

# DART 필드코드 -> 한글 라벨 (주요정보 시트 헤더용; 없으면 코드 그대로)
FIELD_LABELS = {
    "se": "구분", "stock_knd": "주식종류", "thstrm": "당기", "frmtrm": "전기", "lwfr": "전전기",
    "nm": "성명", "sexdstn": "성별", "birth_ym": "출생년월", "ofcps": "직위",
    "rgist_exctv_at": "등기임원여부", "fte_at": "상근여부", "chrg_job": "담당업무",
    "main_career": "주요경력", "mxmm_shrholdr_relate": "최대주주관계",
    "hffc_pd": "재직기간", "tenure_end_on": "임기만료일", "relate": "관계",
    "bsis_posesn_stock_co": "기초주식수", "bsis_posesn_stock_qota_rt": "기초지분율(%)",
    "trmend_posesn_stock_co": "기말주식수", "trmend_posesn_stock_qota_rt": "기말지분율(%)",
    "fo_bbm": "사업부문", "rgllbr_co": "정규직수", "cnttk_co": "계약직수", "sm": "합계",
    "avrg_cnwk_sdytrn": "평균근속연수", "fyer_salary_totamt": "연간급여총액",
    "jan_salary_am": "1인평균급여", "shrholdr_co": "소액주주수", "shrholdr_tot_co": "전체주주수",
    "shrholdr_rate": "소액주주비율(%)", "hold_stock_co": "소액주주보유주식수",
    "stock_tot_co": "총발행주식수", "hold_stock_rate": "소액주주지분율(%)",
    "acqs_mth1": "취득방법1", "acqs_mth2": "취득방법2", "acqs_mth3": "취득방법3",
    "bsis_qy": "기초수량", "change_qy_acqs": "취득수량", "change_qy_dsps": "처분수량",
    "change_qy_incnr": "소각수량", "trmend_qy": "기말수량",
    "isu_dcrs_de": "발행일", "isu_dcrs_stle": "발행형태", "isu_dcrs_stock_knd": "주식종류",
    "isu_dcrs_qy": "수량", "isu_dcrs_mstvdv_fval_amount": "액면가",
    "isu_dcrs_mstvdv_amount": "발행가", "rm": "비고", "stlm_dt": "결산기준일",
}
# 시트에서 숨길 반복 메타 필드
SKIP_FIELDS = {"rcept_no", "corp_cls", "corp_code", "corp_name"}
CELL_MAX = 32000   # 엑셀 셀 최대 문자수(32,767) 안전 여유


def _num(s):
    """DART 금액 문자열 -> 숫자(가능하면 int). 실패 시 원본 문자열 또는 None."""
    if s is None:
        return None
    s = str(s).strip().replace(",", "")
    if s in ("", "-"):
        return None
    try:
        return int(s)
    except ValueError:
        try:
            return float(s)
        except ValueError:
            return s


# ------------------------------------------------------------------ 라우트

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/config")
def api_config():
    # 서버에 키가 설정돼 있으면 프론트에서 키 입력을 선택사항으로 안내.
    return jsonify({
        "has_server_key": bool(os.environ.get("DART_API_KEY", "").strip()),
        "major_groups": MAJOR_GROUPS,
    })


@app.route("/api/toc")
def api_toc():
    """선택 회사의 사업보고서 목차(대분류>소분류)를 반환."""
    key = resolve_key(request.args.get("key"))
    corp_code = request.args.get("corp_code", "")
    year = request.args.get("year", "")
    if not key:
        return jsonify({"ok": False, "error": "DART API 키가 필요합니다."}), 400
    if not corp_code:
        return jsonify({"ok": False, "error": "회사를 먼저 선택하세요."}), 400
    try:
        rcept = dart.find_report_rcept(key, corp_code, year)
        if not rcept:
            return jsonify({"ok": True, "toc": [], "msg": "해당 연도 사업보고서를 찾지 못했습니다."})
        return jsonify({"ok": True, "toc": dart.get_report_toc(key, rcept)})
    except dart.DartError as e:
        return jsonify({"ok": False, "error": f"DART: {e}"}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": f"목차 로드 실패: {e}"}), 500


@app.route("/api/auditor")
def api_auditor():
    """선택 회사들의 외부감사인/감사의견을 반환."""
    key = resolve_key(request.args.get("key"))
    corp_codes = [c for c in request.args.get("corp_codes", "").split(",") if c]
    year = request.args.get("year", "")
    if not key:
        return jsonify({"ok": False, "error": "DART API 키가 필요합니다."}), 400
    if not corp_codes:
        return jsonify({"ok": False, "error": "회사를 선택하세요."}), 400
    try:
        out = []
        for cc in corp_codes:
            rcept = dart.find_report_rcept(key, cc, year)
            rows = dart.get_auditor(key, rcept) if rcept else []
            out.append({"corp_code": cc, "rows": rows})
        return jsonify({"ok": True, "auditors": out})
    except dart.DartError as e:
        return jsonify({"ok": False, "error": f"DART: {e}"}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": f"외부감사인 조회 실패: {e}"}), 500


@app.route("/api/industries")
def api_industries():
    try:
        return jsonify({"ok": True, "industries": naver.get_industries()})
    except Exception as e:
        return jsonify({"ok": False, "error": f"산업군 로드 실패: {e}"}), 500


@app.route("/api/companies")
def api_companies():
    no = request.args.get("no", "")
    key = resolve_key(request.args.get("key"))
    if not no:
        return jsonify({"ok": False, "error": "업종(no)이 필요합니다."}), 400
    try:
        companies = naver.get_companies(no)
        corp_map = dart.get_corp_map(key) if key else {}
        out = []
        for c in companies:
            info = corp_map.get(c["stock_code"])
            out.append({
                "name": c["name"],
                "stock_code": c["stock_code"],
                "corp_code": info["corp_code"] if info else None,
                "in_dart": bool(info),
            })
        return jsonify({"ok": True, "companies": out})
    except dart.DartError as e:
        return jsonify({"ok": False, "error": f"DART: {e}"}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": f"회사 로드 실패: {e}"}), 500


@app.route("/api/accounts")
def api_accounts():
    key = resolve_key(request.args.get("key"))
    corp_codes = [c for c in request.args.get("corp_codes", "").split(",") if c]
    year = request.args.get("year", "")
    reprt = request.args.get("reprt", "11011")
    fs = request.args.get("fs", "CFS")

    if not key:
        return jsonify({"ok": False, "error": "DART API 키가 필요합니다."}), 400
    if not corp_codes:
        return jsonify({"ok": False, "error": "회사를 먼저 선택하세요."}), 400

    try:
        merged, seen = [], set()
        for cc in corp_codes:
            for a in dart.get_accounts(key, cc, year, reprt, fs):
                if a["account_nm"] not in seen:
                    seen.add(a["account_nm"])
                    merged.append(a)
        # 재무제표 표시 순서대로 정렬(여러 회사 합쳐도 유지)
        merged.sort(key=lambda a: a.get("_k") or [99])
        for a in merged:
            a.pop("_k", None)
        return jsonify({"ok": True, "accounts": merged})
    except dart.DartError as e:
        return jsonify({"ok": False, "error": f"DART: {e}"}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": f"계정과목 로드 실패: {e}"}), 500


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _safe_sheet_title(wb, name):
    """엑셀 시트명 제약(31자, 특수문자 금지, 중복불가) 처리."""
    name = re.sub(r"[\[\]:*?/\\]", " ", name).strip()[:31] or "시트"
    base, i, existing = name, 1, set(wb.sheetnames)
    while name in existing:
        i += 1
        suffix = f"_{i}"
        name = base[:31 - len(suffix)] + suffix
    return name


def _write_header_row(ws, row, headers):
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _build_major_sheet(ws, key, companies, years, reprt, endpoint):
    """정기보고서 주요정보 1개 항목을 회사×연도별로 모아 시트 작성."""
    all_rows, cols = [], []   # all_rows: [(company_name, year, row_dict)]
    for c in companies:
        for y in years:
            rows, _ = dart.get_major_info(key, c["corp_code"], y, reprt, endpoint)
            for r in rows:
                all_rows.append((c["name"], y, r))
                if not cols:
                    cols = [k for k in r.keys() if k not in SKIP_FIELDS]
            time.sleep(0.12)   # 연속 호출 제한 완화

    headers = ["회사명", "연도"] + [FIELD_LABELS.get(k, k) for k in cols]
    _write_header_row(ws, 1, headers)
    for ri, (cname, y, r) in enumerate(all_rows, start=2):
        ws.cell(row=ri, column=1, value=cname)
        ws.cell(row=ri, column=2, value=y)
        for ci, k in enumerate(cols, start=3):
            ws.cell(row=ri, column=ci, value=r.get(k))
    if not all_rows:
        ws.cell(row=2, column=1, value="(조회된 데이터 없음)")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 8
    for ci in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = "A2"


def _short_top(top):
    """대분류 제목에서 로마숫자 접두어/괄호 제거 ('II. 사업의 내용' -> '사업의 내용')."""
    s = re.sub(r"^[IVXLC]+\.\s*", "", top)
    s = s.replace("【", "").replace("】", "").strip()
    return s or top


SECTION_SPAN = 14          # 문단 셀 병합 너비(열 수)
SECTION_COLW = 12          # 각 열 너비
_TBL_BORDER = Border(*(4 * [Side(style="thin", color="D0D5DD")]))
_TBL_HDR_FILL = PatternFill("solid", fgColor="EAF0F7")


def _cell_val(s):
    """표 셀 문자열을 가능하면 숫자로. '(1,234)'는 음수로."""
    s = (s or "").strip()
    if not s:
        return None
    t = s.replace(",", "")
    neg = t.startswith("(") and t.endswith(")")
    if neg:
        t = t[1:-1]
    try:
        v = int(t)
    except ValueError:
        try:
            v = float(t)
        except ValueError:
            return s
    return -v if neg else v


def _prose_chunks(text, size=1200):
    return [text[i:i + size] for i in range(0, len(text), size)] or [""]


def _prose_height(text):
    per_line = max(24, SECTION_SPAN * SECTION_COLW // 2)   # 한글 기준 대략
    lines = max(1, -(-len(text) // per_line))
    return min(lines * 15 + 4, 620)


def _build_section_sheet(ws, key, companies, years, top, subs):
    """대분류(top)의 선택 소분류들을 회사×연도별로 '문서형'으로 작성.
    문단은 줄바꿈 문단, 표는 실제 셀 격자로 렌더링."""
    span = SECTION_SPAN
    for ci in range(1, span + 1):
        ws.column_dimensions[get_column_letter(ci)].width = SECTION_COLW
    targets = subs if subs else [top]
    r = 1
    for c in companies:
        for y in years:
            rcept = dart.find_report_rcept(key, c["corp_code"], y)
            for sub in targets:
                # 블록 헤더: 회사 | 연도 | 소분류
                label = f'{c["name"]}    |    {y}    |    {"전체" if sub == top else sub}'
                ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
                hc = ws.cell(row=r, column=1, value=label)
                hc.fill = HEADER_FILL
                hc.font = HEADER_FONT
                hc.alignment = Alignment(vertical="center")
                ws.row_dimensions[r].height = 20
                r += 1

                blocks = dart.get_section_blocks(key, rcept, top, sub) if rcept else []
                if not blocks:
                    ws.cell(row=r, column=1, value="(내용 없음)")
                    r += 2
                    continue

                for kind, payload in blocks:
                    if kind == "text":
                        for chunk in _prose_chunks(payload):
                            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
                            cell = ws.cell(row=r, column=1, value=chunk)
                            cell.alignment = Alignment(wrap_text=True, vertical="top")
                            ws.row_dimensions[r].height = _prose_height(chunk)
                            r += 1
                    else:  # table
                        for ti, trow in enumerate(payload):
                            for ci, val in enumerate(trow[:30], start=1):
                                v = _cell_val(val)
                                cc = ws.cell(row=r, column=ci, value=v)
                                cc.font = Font(size=9, bold=(ti == 0))
                                cc.border = _TBL_BORDER
                                cc.alignment = Alignment(
                                    vertical="top", wrap_text=True,
                                    horizontal="right" if isinstance(v, (int, float)) else "left")
                                if isinstance(v, int):
                                    cc.number_format = "#,##0"
                                if ti == 0:
                                    cc.fill = _TBL_HDR_FILL
                            r += 1
                    r += 1   # 블록 사이 빈 줄
    ws.freeze_panes = "A2"


def _build_accounts_sheet(ws, key, companies, years, reprt, fs, accounts, industry):
    """계정과목 비교 시트. 열 = 회사(병합) × 연도, 행 = 계정과목."""
    per = len(years)
    col_specs = []   # [(company_name, year, {account_nm: amount})]  회사-major 순서
    for c in companies:
        for y in years:
            rows = dart.get_statement(key, c["corp_code"], y, reprt, fs)
            amap = {}
            for r in rows:
                nm = (r.get("account_nm") or "").strip()
                if nm and nm not in amap:
                    amap[nm] = _num(r.get("thstrm_amount"))
            col_specs.append((c["name"], y, amap))

    meta = [
        "재무분석기 · CPA KOOK",
        f"산업군: {industry}",
        f"연도: {', '.join(str(y) for y in years)}   보고서: {REPRT_NAMES.get(reprt, reprt)}   구분: {FS_NAMES.get(fs, fs)}",
        f"생성일시: {datetime.now():%Y-%m-%d %H:%M}   |   작성: CPA KOOK",
    ]
    for i, line in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=line).font = Font(bold=(i == 1), size=13 if i == 1 else 10)

    h1 = len(meta) + 2   # 회사명 헤더(병합)
    h2 = h1 + 1          # 연도 헤더

    # 계정과목 헤더(2행 병합)
    ws.merge_cells(start_row=h1, start_column=1, end_row=h2, end_column=1)
    cell = ws.cell(row=h1, column=1, value="계정과목")
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")

    # 회사(병합) + 연도 헤더
    for ci, c in enumerate(companies):
        start = 2 + ci * per
        for yi, y in enumerate(years):
            yc = ws.cell(row=h2, column=start + yi, value=y)
            yc.fill = HEADER_FILL; yc.font = HEADER_FONT
            yc.alignment = Alignment(horizontal="center")
        if per > 1:
            ws.merge_cells(start_row=h1, start_column=start, end_row=h1, end_column=start + per - 1)
        nc = ws.cell(row=h1, column=start, value=c["name"])
        nc.fill = HEADER_FILL; nc.font = HEADER_FONT
        nc.alignment = Alignment(horizontal="center")

    # 본문
    for i, acc in enumerate(accounts):
        r = h2 + 1 + i
        ws.cell(row=r, column=1, value=acc)
        for j, (_, _, amap) in enumerate(col_specs, start=2):
            v = amap.get(acc)
            cell = ws.cell(row=r, column=j, value=v)
            if isinstance(v, (int, float)):
                cell.number_format = "#,##0"

    ws.column_dimensions["A"].width = 28
    for j in range(2, len(col_specs) + 2):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.freeze_panes = ws.cell(row=h2 + 1, column=2)


def _metric_cellval(v, kind):
    if v is None:
        return None
    return v / 1_000_000 if kind == "amount" else v   # 금액은 백만원


def _metric_numfmt(kind):
    return {"amount": "#,##0", "pct": "0.0%", "days": '0"일"'}.get(kind, "0.00")


def _change_numfmt(kind):
    """전년比 서식: 금액=증감률(%), 비율=%p, 회전=배수, 일수=일."""
    if kind in ("amount", "pct"):
        return "+0.0%;-0.0%"
    if kind == "days":
        return '+0"일";-0"일"'
    return "+0.00;-0.00"


def _build_metrics_sheet(ws, key, companies, years, reprt, fs):
    """반도체 분석 세트: 회사×연도 지표 자동계산 + 연도마다 전년比."""
    ncol = 1
    r = 1
    for c in companies:
        data = metrics.compute_metrics(key, c["corp_code"], years, reprt, fs)
        ys = data["years"]
        # 헤더: 지표 | y0 | y1 | 전년比 | y2 | 전년比 ...
        headers, change_cols = ["지표"], set()
        for i, y in enumerate(ys):
            headers.append(str(y))
            if i > 0:
                headers.append("전년比")
                change_cols.add(len(headers))   # 1-based 열 번호
        ncol = len(headers)

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
        hc = ws.cell(row=r, column=1,
                     value=f'{c["name"]}  ·  분석지표  (금액 백만원 · 전년比: 금액=증감률, 비율=%p차이)')
        hc.fill = HEADER_FILL
        hc.font = HEADER_FONT
        hc.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 20
        r += 1

        _write_header_row(ws, r, headers)
        r += 1

        for row in data["rows"]:
            if row["kind"] == "sep":
                r += 1
                continue
            kind = row["kind"]
            ws.cell(row=r, column=1, value=row["label"])
            col = 2
            for i, y in enumerate(ys):
                cell = ws.cell(row=r, column=col, value=_metric_cellval(row["values"].get(y), kind))
                cell.number_format = _metric_numfmt(kind)
                col += 1
                if i > 0:                          # 이 연도의 전년比
                    ch = ws.cell(row=r, column=col, value=row["changes"].get(y))
                    ch.number_format = _change_numfmt(kind)
                    ch.font = Font(size=10, color="6B7684")
                    col += 1
            r += 1
        r += 1   # 회사 간 빈 줄

    ws.column_dimensions["A"].width = 26
    for j in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(j)].width = 11 if j in change_cols else 14
    ws.freeze_panes = "B1"


@app.route("/api/export", methods=["POST"])
def api_export():
    data = request.get_json(force=True)
    key = resolve_key(data.get("key"))
    # 다중 연도(최대 5). 과거 호환: year 단일값도 허용.
    years = data.get("years") or ([data.get("year")] if data.get("year") else [])
    years = [str(y) for y in years if str(y).strip()][:5]
    reprt = data.get("reprt", "11011")
    fs = data.get("fs", "CFS")
    industry = data.get("industry", "")
    companies = data.get("companies", [])   # [{name, corp_code}]
    accounts = data.get("accounts", [])     # [account_nm, ...]
    metrics_on = bool(data.get("metrics"))   # 분석지표(반도체 세트) 시트 포함 여부
    # 플러스알파(선택): 사업보고서 내용
    major = [m for m in data.get("major", []) if m in MAJOR_MAP]   # 주요정보 endpoint id
    # 본문 섹션: [{top, sub}] -> 대분류별로 소분류 묶기
    sec_by_top = OrderedDict()
    for it in data.get("sections", []):
        top = (it or {}).get("top")
        sub = (it or {}).get("sub")
        if not top:
            continue
        sec_by_top.setdefault(top, [])
        if sub and sub != top and sub not in sec_by_top[top]:
            sec_by_top[top].append(sub)
    sections = sec_by_top

    if not (key and companies):
        return jsonify({"ok": False, "error": "API 키와 회사를 선택하세요."}), 400
    if not years:
        return jsonify({"ok": False, "error": "연도를 1개 이상 선택하세요."}), 400
    if not (metrics_on or accounts or major or sections):
        return jsonify({"ok": False, "error": "분석지표·계정과목·사업보고서 내용 중 하나 이상 선택하세요."}), 400

    try:
        wb = Workbook()
        first_used = False

        # ---- 0) 분석지표 시트 (반도체 세트) -----------------------------
        if metrics_on:
            ws = wb.active
            ws.title = "분석지표"
            first_used = True
            _build_metrics_sheet(ws, key, companies, years, reprt, fs)

        # ---- 1) 계정과목 시트 (선택 시) --------------------------------
        if accounts:
            ws = wb.active if not first_used else wb.create_sheet()
            ws.title = "재무분석"
            first_used = True
            _build_accounts_sheet(ws, key, companies, years, reprt, fs, accounts, industry)

        # ---- 2) 정기보고서 주요정보 시트들 -----------------------------
        for mid in major:
            ws = wb.active if not first_used else wb.create_sheet()
            ws.title = _safe_sheet_title(wb, MAJOR_MAP[mid])
            first_used = True
            _build_major_sheet(ws, key, companies, years, reprt, mid)

        # ---- 3) 사업보고서 본문 섹션 시트들 (대분류별 시트) -------------
        for top, subs in sections.items():
            ws = wb.active if not first_used else wb.create_sheet()
            ws.title = _safe_sheet_title(wb, _short_top(top))
            first_used = True
            _build_section_sheet(ws, key, companies, years, top, subs)

    except dart.DartError as e:
        return jsonify({"ok": False, "error": f"DART: {e}"}), 400
    except Exception as e:
        return jsonify({"ok": False, "error": f"데이터 수집 실패: {e}"}), 500

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    yr_tag = "-".join(str(y) for y in years)
    fname = f"재무분석_{industry}_{yr_tag}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    return send_file(
        bio,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    import webbrowser, threading
    url = "http://127.0.0.1:5000"
    threading.Timer(1.2, lambda: webbrowser.open(url)).start()
    print(f"\n  재무분석기 실행 중 -> {url}\n  (종료하려면 이 창에서 Ctrl+C)\n")
    app.run(host="127.0.0.1", port=5000, debug=False)
